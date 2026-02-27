"""
主要页面路由
"""

from flask import Blueprint, render_template, jsonify, request
from flask_login import login_required, current_user
import pandas as pd
import os
import json
from datetime import datetime
from .models import Upload, db, PPTTemplate, PPTProject, MeetingMinutes, TodoItem
from .data_cleaner import DataCleaner
from .ai_analyzer import AIAnalyzer
from .ppt_manager import PPTManager
from .meeting_minutes import MeetingMinutesAssistant

bp = Blueprint('main', __name__)

@bp.route('/')
def index():
    """首页"""
    return render_template('main/index.html')

@bp.route('/dashboard')
@login_required
def dashboard():
    """用户仪表板"""
    return render_template('main/dashboard.html')

@bp.route('/api/status')
def api_status():
    """API状态检查"""
    return jsonify({
        'success': True,
        'service': '智能办公文档处理助手',
        'version': '1.0.0',
        'status': 'running'
    })

@bp.route('/api/user/info')
@login_required
def user_info():
    """获取当前用户信息"""
    return jsonify({
        'success': True,
        'data': {
            'id': current_user.id,
            'username': current_user.username,
            'email': current_user.email,
            'created_at': current_user.created_at.isoformat() if current_user.created_at else None
        }
    })

@bp.route('/api/excel/preview/<int:file_id>')
@login_required
def excel_preview(file_id):
    """Excel数据预览API - 增强版本
    返回前10行数据和统计信息，格式为 {
        "data": [行数据列表],
        "stats": {
            "row_count": 行数,
            "column_count": 列数,
            "data_types": {数据类型: 数量},
            "missing_values_total": 缺失值总数,
            "missing_values_by_column": {列名: 缺失数},
            "numeric_stats": {列名: {统计量: 值}} (仅数值列),
            "preprocessing": {预处理信息}
        }
    }
    """
    # 获取上传记录
    upload = Upload.query.filter_by(id=file_id, user_id=current_user.id).first()
    
    if not upload:
        return jsonify({'success': False, 'message': '文件不存在或无权访问'}), 404
    
    # 检查文件是否存在
    if not os.path.exists(upload.upload_path):
        return jsonify({'success': False, 'message': '文件不存在于服务器'}), 404
    
    # 检查文件扩展名
    _, ext = os.path.splitext(upload.filename)
    ext = ext.lower()
    
    # 支持的文件格式
    supported_formats = ['.csv', '.xlsx', '.xls']
    if ext not in supported_formats:
        return jsonify({
            'success': False, 
            'message': f'不支持的文件格式: {ext}。支持格式: {", ".join(supported_formats)}'
        }), 400
    
    try:
        # 读取文件
        if ext == '.csv':
            df = pd.read_csv(upload.upload_path)
        else:
            # 使用openpyxl引擎读取Excel文件，并处理合并单元格
            import openpyxl
            from openpyxl.utils import get_column_letter
            
            # 加载工作簿
            wb = openpyxl.load_workbook(upload.upload_path, data_only=True)
            ws = wb.active
            
            # 获取所有合并单元格范围
            merged_ranges = ws.merged_cells.ranges
            
            # 展开合并单元格：将合并区域左上角的值复制到所有单元格
            for merged_range in list(merged_ranges):
                # 获取合并区域的左上角单元格值
                top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                top_left_value = top_left_cell.value
                
                # 将值复制到合并区域的所有单元格
                for row in range(merged_range.min_row, merged_range.max_row + 1):
                    for col in range(merged_range.min_col, merged_range.max_col + 1):
                        cell = ws.cell(row=row, column=col)
                        cell.value = top_left_value
            
            # 将工作表转换为DataFrame
            data = ws.values
            columns = next(data)  # 第一行为列名
            df = pd.DataFrame(data, columns=columns)
        
        # 记录原始形状
        original_shape = df.shape
        
        # 数据预处理：删除完全空的行和列
        df = df.dropna(how='all')
        df = df.dropna(axis=1, how='all')
        
        # 获取前10行数据（转换为JSON可序列化格式）
        preview_data = df.head(10).replace({pd.NaT: None, pd.NaN: None}).to_dict(orient='records')
        
        # 计算统计信息
        # 1. 基本维度
        row_count = len(df)
        column_count = len(df.columns)
        
        # 2. 数据类型分布
        dtypes_count = df.dtypes.value_counts()
        data_types = {}
        for dtype, count in dtypes_count.items():
            dtype_str = str(dtype)
            data_types[dtype_str] = int(count)
        
        # 3. 缺失值统计
        missing_total = int(df.isnull().sum().sum())
        missing_by_column = {}
        for col, count in df.isnull().sum().items():
            missing_by_column[col] = int(count)
        
        # 4. 数值列统计
        numeric_cols = df.select_dtypes(include=['number']).columns
        numeric_stats = {}
        if len(numeric_cols) > 0:
            numeric_desc = df[numeric_cols].describe()
            for col in numeric_desc.columns:
                numeric_stats[col] = {}
                for stat, value in numeric_desc[col].items():
                    if pd.isna(value):
                        numeric_stats[col][stat] = None
                    else:
                        # 转换为Python原生类型
                        try:
                            # 尝试转换为float，如果失败则转为字符串
                            numeric_stats[col][stat] = float(value)
                        except (ValueError, TypeError):
                            numeric_stats[col][stat] = str(value)
        
        # 5. 数据预处理信息
        preprocessing_info = {
            'removed_empty_rows': original_shape[0] - row_count,
            'removed_empty_columns': original_shape[1] - column_count,
            'original_shape': list(original_shape),
            'cleaned_shape': [row_count, column_count]
        }
        
        stats = {
            'row_count': row_count,
            'column_count': column_count,
            'data_types': data_types,
            'missing_values_total': missing_total,
            'missing_values_by_column': missing_by_column,
            'numeric_stats': numeric_stats,
            'preprocessing': preprocessing_info
        }
        
        return jsonify({
            'success': True,
            'data': preview_data,
            'stats': stats
        })
        
    except pd.errors.EmptyDataError:
        return jsonify({'success': False, 'message': '文件内容为空，无法读取'}), 400
    except pd.errors.ParserError as e:
        return jsonify({'success': False, 'message': f'文件解析错误：{str(e)}'}), 400
    except pd.errors.UnsupportedFileTypeError as e:
        return jsonify({'success': False, 'message': f'不支持的文件类型：{str(e)}'}), 400
    except pd.errors.DataError as e:
        return jsonify({'success': False, 'message': f'数据错误：{str(e)}'}), 400
    except FileNotFoundError:
        return jsonify({'success': False, 'message': '文件不存在'}), 404
    except PermissionError:
        return jsonify({'success': False, 'message': '没有文件读取权限'}), 403
    except pd.errors.OutOfBoundsDatetime:
        return jsonify({'success': False, 'message': '日期时间超出范围，请检查日期格式'}), 400
    except Exception as e:
        # 通用异常捕获，提供更详细的错误信息
        error_type = type(e).__name__
        return jsonify({
            'success': False, 
            'message': f'文件读取失败：{error_type}: {str(e)}'
        }), 500

@bp.route('/excel-cleaner')
@login_required
def excel_cleaner_page():
    """Excel数据清洗页面"""
    return render_template('main/excel_cleaner.html')

@bp.route('/api/user/uploads')
@login_required
def user_uploads():
    """获取用户上传文件列表"""
    try:
        # 获取用户的所有上传记录，按时间倒序排列
        uploads = Upload.query.filter_by(user_id=current_user.id)\
            .order_by(Upload.uploaded_at.desc())\
            .all()
        
        uploads_data = []
        for upload in uploads:
            uploads_data.append({
                'id': upload.id,
                'filename': upload.filename,
                'original_filename': upload.original_filename,
                'file_size': upload.file_size,
                'file_type': upload.file_type,
                'uploaded_at': upload.uploaded_at.isoformat() if upload.uploaded_at else None
            })
        
        return jsonify({
            'success': True,
            'data': uploads_data,
            'count': len(uploads_data)
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'获取上传列表失败: {str(e)}'
        }), 500

@bp.route('/api/excel/clean/<int:file_id>', methods=['POST'])
@login_required
def excel_clean(file_id):
    """Excel数据清洗API
    支持多种清洗选项，返回清洗后的数据和详细报告
    """
    # 获取上传记录
    upload = Upload.query.filter_by(id=file_id, user_id=current_user.id).first()
    
    if not upload:
        return jsonify({'success': False, 'message': '文件不存在或无权访问'}), 404
    
    # 检查文件是否存在
    if not os.path.exists(upload.upload_path):
        return jsonify({'success': False, 'message': '文件不存在于服务器'}), 404
    
    # 获取清洗选项
    try:
        options = request.get_json()
        if not options:
            options = {}
    except:
        options = {}
    
    # 设置默认选项
    default_options = {
        'missing_strategy': 'drop',
        'remove_duplicates': True,
        'type_conversions': {},
        'outlier_strategy': None
    }
    
    # 合并选项
    for key, value in default_options.items():
        if key not in options:
            options[key] = value
    
    try:
        # 创建数据清洗器
        cleaner = DataCleaner(upload.upload_path)
        
        # 加载数据
        if not cleaner.load_data():
            return jsonify({'success': False, 'message': '数据加载失败'}), 400
        
        # 分析数据质量
        quality_report = cleaner.analyze_data_quality()
        
        # 清洗数据
        cleaned_df, cleaning_report = cleaner.clean_data(options)
        
        # 获取清洗后的预览数据
        preview_data = cleaned_df.head(10).replace({pd.NaT: None, pd.NaN: None}).to_dict(orient='records')
        
        # 计算清洗后统计信息
        cleaned_stats = {
            'row_count': len(cleaned_df),
            'column_count': len(cleaned_df.columns),
            'data_types': {str(dtype): int(count) for dtype, count in cleaned_df.dtypes.value_counts().items()},
            'missing_values_total': int(cleaned_df.isnull().sum().sum()),
            'missing_values_by_column': {col: int(count) for col, count in cleaned_df.isnull().sum().items()}
        }
        
        return jsonify({
            'success': True,
            'data': preview_data,
            'stats': cleaned_stats,
            'quality_report': quality_report,
            'cleaning_report': cleaning_report,
            'applied_options': options
        })
        
    except Exception as e:
        error_type = type(e).__name__
        return jsonify({
            'success': False,
            'message': f'数据清洗失败：{error_type}: {str(e)}'
        }), 500

@bp.route('/api/excel/ai-analyze/<int:file_id>')
@login_required
def excel_ai_analyze(file_id):
    """Excel数据AI分析API
    使用DeepSeek API进行数据质量分析和智能建议生成
    """
    # 获取上传记录
    upload = Upload.query.filter_by(id=file_id, user_id=current_user.id).first()
    
    if not upload:
        return jsonify({'success': False, 'message': '文件不存在或无权访问'}), 404
    
    # 检查文件是否存在
    if not os.path.exists(upload.upload_path):
        return jsonify({'success': False, 'message': '文件不存在于服务器'}), 404
    
    try:
        # 创建数据清洗器
        cleaner = DataCleaner(upload.upload_path)
        
        # 加载数据
        if not cleaner.load_data():
            return jsonify({'success': False, 'message': '数据加载失败'}), 400
        
        # 分析数据质量
        quality_report = cleaner.analyze_data_quality()
        
        # 获取数据样本（前10行）
        cleaner.load_data()  # 确保数据已加载
        sample_data = cleaner.df.head(10).replace({pd.NaT: None, pd.NaN: None}).to_dict(orient='records')
        
        # 创建AI分析器
        ai_analyzer = AIAnalyzer()
        
        # 进行AI分析
        ai_report = ai_analyzer.analyze_data_quality(quality_report, sample_data)
        
        return jsonify({
            'success': True,
            'quality_report': quality_report,
            'ai_analysis': ai_report,
            'sample_data': sample_data
        })
        
    except Exception as e:
        error_type = type(e).__name__
        return jsonify({
            'success': False,
            'message': f'AI分析失败：{error_type}: {str(e)}'
        }), 500

@bp.route('/api/excel/export/<int:file_id>')
@login_required
def excel_export(file_id):
    """Excel数据导出API
    支持多种格式导出清洗后的数据
    """
    # 获取上传记录
    upload = Upload.query.filter_by(id=file_id, user_id=current_user.id).first()
    
    if not upload:
        return jsonify({'success': False, 'message': '文件不存在或无权访问'}), 404
    
    # 检查文件是否存在
    if not os.path.exists(upload.upload_path):
        return jsonify({'success': False, 'message': '文件不存在于服务器'}), 404
    
    # 获取导出参数
    format_type = request.args.get('format', 'csv').lower()
    options = request.args.get('options', '{}')
    
    try:
        options_dict = json.loads(options)
    except:
        options_dict = {}
    
    # 支持的格式
    supported_formats = ['csv', 'excel', 'json', 'html']
    if format_type not in supported_formats:
        return jsonify({
            'success': False,
            'message': f'不支持的导出格式: {format_type}。支持格式: {", ".join(supported_formats)}'
        }), 400
    
    try:
        # 创建数据清洗器
        cleaner = DataCleaner(upload.upload_path)
        
        # 加载数据
        if not cleaner.load_data():
            return jsonify({'success': False, 'message': '数据加载失败'}), 400
        
        # 使用默认选项清洗数据（确保有清洗后的数据）
        default_options = {
            'missing_strategy': 'drop',
            'remove_duplicates': True,
            'type_conversions': {},
            'outlier_strategy': None
        }
        
        # 合并用户选项
        for key, value in options_dict.items():
            default_options[key] = value
        
        cleaned_df, _ = cleaner.clean_data(default_options)
        
        # 生成导出文件路径
        base_name = os.path.splitext(upload.original_filename)[0]
        export_dir = os.path.join(os.path.dirname(upload.upload_path), "exports")
        os.makedirs(export_dir, exist_ok=True)
        
        export_filename = f"{base_name}_cleaned.{format_type}"
        export_path = os.path.join(export_dir, export_filename)
        
        # 根据格式导出
        if format_type == 'csv':
            cleaned_df.to_csv(export_path, index=False, encoding='utf-8-sig')
        elif format_type == 'excel':
            cleaned_df.to_excel(export_path, index=False)
        elif format_type == 'json':
            cleaned_df.to_json(export_path, orient='records', indent=2)
        elif format_type == 'html':
            cleaned_df.to_html(export_path, index=False)
        
        # 返回文件下载信息
        return jsonify({
            'success': True,
            'message': '数据导出成功',
            'export_file': export_filename,
            'format': format_type,
            'download_url': f"/static/uploads/exports/{export_filename}"
        })
        
    except Exception as e:
        error_type = type(e).__name__
        return jsonify({
            'success': False,
            'message': f'数据导出失败：{error_type}: {str(e)}'
        }), 500

# ==================== PPT生成相关路由 ====================

@bp.route('/ppt-generator')
@login_required
def ppt_generator_page():
    """PPT生成器页面"""
    return render_template('main/ppt_generator.html')

@bp.route('/api/ppt/templates')
@login_required
def ppt_templates_list():
    """获取PPT模板列表"""
    try:
        category = request.args.get('category')
        style_type = request.args.get('style_type')
        public_only = request.args.get('public_only', 'false').lower() == 'true'
        
        ppt_manager = PPTManager()
        templates = ppt_manager.list_templates(current_user.id, category, style_type, public_only)
        
        templates_data = []
        for template in templates:
            templates_data.append({
                'id': template.id,
                'name': template.name,
                'description': template.description,
                'category': template.category,
                'thumbnail_url': f'/static/{template.thumbnail_path}' if template.thumbnail_path else None,
                'style_type': template.style_type,
                'tags': template.tags.split(',') if template.tags else [],
                'is_public': template.is_public,
                'created_at': template.created_at.isoformat() if template.created_at else None
            })
        
        return jsonify({
            'success': True,
            'data': templates_data,
            'count': len(templates_data)
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'获取模板列表失败: {str(e)}'
        }), 500

@bp.route('/api/ppt/templates/upload', methods=['POST'])
@login_required
def ppt_template_upload():
    """上传PPT模板"""
    try:
        if 'template_file' not in request.files:
            return jsonify({'success': False, 'message': '请选择模板文件'}), 400
        
        template_file = request.files['template_file']
        if template_file.filename == '':
            return jsonify({'success': False, 'message': '请选择模板文件'}), 400
        
        name = request.form.get('name', '未命名模板')
        description = request.form.get('description', '')
        category = request.form.get('category', '商务')
        style_type = request.form.get('style_type', '信息图风')
        tags = request.form.get('tags', '').split(',') if request.form.get('tags') else None
        is_public = request.form.get('is_public', 'false').lower() == 'true'
        
        ppt_manager = PPTManager()
        template, message = ppt_manager.upload_template(
            current_user.id,
            template_file,
            name,
            description,
            category,
            style_type,
            tags,
            is_public
        )
        
        if template:
            return jsonify({
                'success': True,
                'message': message,
                'data': {
                    'id': template.id,
                    'name': template.name
                }
            })
        else:
            return jsonify({'success': False, 'message': message}), 400
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'模板上传失败: {str(e)}'
        }), 500

@bp.route('/api/ppt/templates/<int:template_id>', methods=['DELETE'])
@login_required
def ppt_template_delete(template_id):
    """删除PPT模板"""
    try:
        ppt_manager = PPTManager()
        success, message = ppt_manager.delete_template(template_id, current_user.id)
        
        if success:
            return jsonify({'success': True, 'message': message})
        else:
            return jsonify({'success': False, 'message': message}), 400
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'模板删除失败: {str(e)}'
        }), 500

@bp.route('/api/ppt/projects', methods=['POST'])
@login_required
def ppt_project_create():
    """创建PPT项目"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': '请求数据为空'}), 400
        
        title = data.get('title', '未命名项目')
        description = data.get('description', '')
        template_id = data.get('template_id')
        content_data = data.get('content_data')
        
        ppt_manager = PPTManager()
        project, message = ppt_manager.create_project(
            current_user.id,
            title,
            description,
            template_id,
            content_data
        )
        
        if project:
            return jsonify({
                'success': True,
                'message': message,
                'data': {
                    'id': project.id,
                    'title': project.title
                }
            })
        else:
            return jsonify({'success': False, 'message': message}), 400
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'项目创建失败: {str(e)}'
        }), 500

@bp.route('/api/ppt/projects/<int:project_id>/generate-pptx', methods=['POST'])
@login_required
def ppt_generate_pptx(project_id):
    """生成PPTX文件"""
    try:
        ppt_manager = PPTManager()
        success, message, file_path = ppt_manager.generate_pptx(project_id, current_user.id)
        
        if success:
            return jsonify({
                'success': True,
                'message': message,
                'data': {
                    'file_path': file_path,
                    'download_url': f'/static/{file_path}' if file_path else None
                }
            })
        else:
            return jsonify({'success': False, 'message': message}), 400
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'PPT生成失败: {str(e)}'
        }), 500

@bp.route('/api/ppt/projects/<int:project_id>/generate-html', methods=['POST'])
@login_required
def ppt_generate_html(project_id):
    """生成HTML版本"""
    try:
        ppt_manager = PPTManager()
        success, message, file_path = ppt_manager.generate_html(project_id, current_user.id)
        
        if success:
            return jsonify({
                'success': True,
                'message': message,
                'data': {
                    'file_path': file_path,
                    'view_url': f'/static/{file_path}' if file_path else None
                }
            })
        else:
            return jsonify({'success': False, 'message': message}), 400
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'HTML生成失败: {str(e)}'
        }), 500

@bp.route('/api/ppt/projects/<int:project_id>/share', methods=['POST'])
@login_required
def ppt_create_share_link(project_id):
    """创建分享链接"""
    try:
        data = request.get_json()
        expires_hours = data.get('expires_hours', 24) if data else 24
        
        ppt_manager = PPTManager()
        success, message, share_url = ppt_manager.create_share_link(
            project_id,
            current_user.id,
            expires_hours
        )
        
        if success:
            return jsonify({
                'success': True,
                'message': message,
                'data': {
                    'share_url': share_url,
                    'full_url': f'{request.host_url.rstrip("/")}{share_url}'
                }
            })
        else:
            return jsonify({'success': False, 'message': message}), 400
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'分享链接创建失败: {str(e)}'
        }), 500

@bp.route('/api/ppt/optimize', methods=['POST'])
@login_required
def ppt_optimize_content():
    """AI优化PPT内容"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': '请求数据为空'}), 400
        
        content_data = data.get('content_data')
        optimization_type = data.get('optimization_type', 'clarity')
        
        if not content_data:
            return jsonify({'success': False, 'message': '内容数据不能为空'}), 400
        
        ppt_manager = PPTManager()
        optimized_content, message = ppt_manager.optimize_content_with_ai(
            content_data,
            optimization_type
        )
        
        return jsonify({
            'success': True,
            'message': message,
            'data': {
                'optimized_content': optimized_content
            }
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'内容优化失败: {str(e)}'
        }), 500

@bp.route('/api/ppt/search-images', methods=['POST'])
@login_required
def ppt_search_images():
    """搜索匹配图片"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': '请求数据为空'}), 400
        
        keywords = data.get('keywords', [])
        count = data.get('count', 5)
        
        if not keywords:
            return jsonify({'success': False, 'message': '关键词不能为空'}), 400
        
        ppt_manager = PPTManager()
        images, message = ppt_manager.search_unsplash_images(keywords, count)
        
        return jsonify({
            'success': True,
            'message': message,
            'data': {
                'images': images
            }
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'图片搜索失败: {str(e)}'
        }), 500

@bp.route('/api/ppt/match-images', methods=['POST'])
@login_required
def ppt_match_images():
    """为内容匹配合适图片"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': '请求数据为空'}), 400
        
        content_data = data.get('content_data')
        image_count = data.get('image_count', 3)
        
        if not content_data:
            return jsonify({'success': False, 'message': '内容数据不能为空'}), 400
        
        ppt_manager = PPTManager()
        matched_images = ppt_manager.match_images_to_content(content_data, image_count)
        
        return jsonify({
            'success': True,
            'message': '图片匹配成功',
            'data': {
                'images': matched_images
            }
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'图片匹配失败: {str(e)}'
        }), 500

@bp.route('/ppt/share/<share_token>')
def ppt_share_view(share_token):
    """PPT分享页面"""
    # 查找项目
    project = PPTProject.query.filter_by(share_token=share_token).first()
    
    if not project:
        return render_template('errors/404.html'), 404
    
    # 检查是否过期
    if project.share_expires and project.share_expires < datetime.utcnow():
        return render_template('errors/expired.html'), 410
    
    # 渲染分享页面
    return render_template('main/ppt_share.html', project=project)

@bp.route('/api/ppt/projects/<int:project_id>')
@login_required
def ppt_project_detail(project_id):
    """获取PPT项目详情"""
    try:
        project = PPTProject.query.filter_by(id=project_id, user_id=current_user.id).first()
        
        if not project:
            return jsonify({'success': False, 'message': '项目不存在或无权访问'}), 404
        
        project_data = {
            'id': project.id,
            'title': project.title,
            'description': project.description,
            'template_id': project.template_id,
            'content_data': json.loads(project.content_data) if project.content_data else {},
            'generated_pptx_path': project.generated_pptx_path,
            'generated_html_path': project.generated_html_path,
            'share_token': project.share_token,
            'share_expires': project.share_expires.isoformat() if project.share_expires else None,
            'status': project.status,
            'created_at': project.created_at.isoformat() if project.created_at else None,
            'updated_at': project.updated_at.isoformat() if project.updated_at else None
        }
        
        return jsonify({
            'success': True,
            'data': project_data
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'获取项目详情失败: {str(e)}'
        }), 500

@bp.route('/api/ppt/projects')
@login_required
def ppt_projects_list():
    """获取用户PPT项目列表"""
    try:
        projects = PPTProject.query.filter_by(user_id=current_user.id)\
            .order_by(PPTProject.updated_at.desc())\
            .all()
        
        projects_data = []
        for project in projects:
            projects_data.append({
                'id': project.id,
                'title': project.title,
                'description': project.description,
                'status': project.status,
                'created_at': project.created_at.isoformat() if project.created_at else None,
                'updated_at': project.updated_at.isoformat() if project.updated_at else None
            })
        
        return jsonify({
            'success': True,
            'data': projects_data,
            'count': len(projects_data)
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'获取项目列表失败: {str(e)}'
        }), 500

# ==================== 会议纪要助手相关路由 ====================

@bp.route('/meeting-minutes')
@login_required
def meeting_minutes_page():
    """会议纪要助手页面"""
    return render_template('main/meeting_minutes.html')

@bp.route('/api/meeting-minutes/process', methods=['POST'])
@login_required
def process_meeting_minutes():
    """处理会议文本文件"""
    try:
        # 获取文件ID
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': '请求数据为空'}), 400
        
        file_id = data.get('file_id')
        language = data.get('language', 'zh')
        
        if not file_id:
            return jsonify({'success': False, 'message': '文件ID不能为空'}), 400
        
        # 获取上传记录
        upload = Upload.query.filter_by(id=file_id, user_id=current_user.id).first()
        
        if not upload:
            return jsonify({'success': False, 'message': '文件不存在或无权访问'}), 404
        
        # 检查文件是否存在
        if not os.path.exists(upload.upload_path):
            return jsonify({'success': False, 'message': '文件不存在于服务器'}), 404
        
        # 检查文件类型
        _, ext = os.path.splitext(upload.filename)
        ext = ext.lower()
        
        if ext not in ['.txt', '.docx']:
            return jsonify({
                'success': False, 
                'message': f'不支持的文件格式: {ext}。仅支持.txt和.docx'
            }), 400
        
        # 处理会议文本
        assistant = MeetingMinutesAssistant()
        result = assistant.process_meeting_text(upload.upload_path, language)
        
        # 保存到数据库
        meeting_minutes = MeetingMinutes(
            title=f"会议纪要_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
            original_file_id=file_id,
            original_text=result['text_processing']['raw_text_sample'],
            summary=result['summary']['summary_text'],
            structured_data=json.dumps(result['summary']['structured_data'], ensure_ascii=False),
            todo_items=json.dumps(result['todo_items'], ensure_ascii=False),
            timeline_data=json.dumps(result['timeline']['data'], ensure_ascii=False),
            language=language,
            processing_status='completed',
            user_id=current_user.id
        )
        
        db.session.add(meeting_minutes)
        db.session.commit()
        
        # 保存待办事项到独立表
        for todo_item in result['todo_items']:
            todo = TodoItem(
                meeting_minutes_id=meeting_minutes.id,
                description=todo_item['description'],
                assignee=todo_item.get('assignee', ''),
                priority=todo_item.get('priority', 2),
                due_date=todo_item.get('due_date'),
                status='pending',
                user_id=current_user.id
            )
            db.session.add(todo)
        
        db.session.commit()
        
        # 返回结果
        return jsonify({
            'success': True,
            'message': '会议纪要处理完成',
            'data': {
                'meeting_minutes_id': meeting_minutes.id,
                'title': meeting_minutes.title,
                'summary': result['summary']['summary_text'],
                'todo_items': result['todo_items'],
                'timeline_data': result['timeline']['data'],
                'chart_config': result['timeline']['chart_config'],
                'quality_report': result['text_processing']['quality_report']
            }
        })
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"会议纪要处理失败: {e}")
        return jsonify({
            'success': False,
            'message': f'会议纪要处理失败: {str(e)}'
        }), 500

@bp.route('/api/meeting-minutes/list')
@login_required
def meeting_minutes_list():
    """获取用户会议纪要列表"""
    try:
        minutes_list = MeetingMinutes.query.filter_by(user_id=current_user.id)\
            .order_by(MeetingMinutes.created_at.desc())\
            .all()
        
        data = []
        for minutes in minutes_list:
            data.append({
                'id': minutes.id,
                'title': minutes.title,
                'language': minutes.language,
                'processing_status': minutes.processing_status,
                'created_at': minutes.created_at.isoformat() if minutes.created_at else None,
                'updated_at': minutes.updated_at.isoformat() if minutes.updated_at else None
            })
        
        return jsonify({
            'success': True,
            'data': data,
            'count': len(data)
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'获取会议纪要列表失败: {str(e)}'
        }), 500

@bp.route('/api/meeting-minutes/<int:minutes_id>')
@login_required
def meeting_minutes_detail(minutes_id):
    """获取会议纪要详情"""
    try:
        minutes = MeetingMinutes.query.filter_by(id=minutes_id, user_id=current_user.id).first()
        
        if not minutes:
            return jsonify({'success': False, 'message': '会议纪要不存在或无权访问'}), 404
        
        # 解析JSON字段
        structured_data = json.loads(minutes.structured_data) if minutes.structured_data else {}
        todo_items = json.loads(minutes.todo_items) if minutes.todo_items else []
        timeline_data = json.loads(minutes.timeline_data) if minutes.timeline_data else []
        
        data = {
            'id': minutes.id,
            'title': minutes.title,
            'original_file_id': minutes.original_file_id,
            'original_text': minutes.original_text,
            'summary': minutes.summary,
            'structured_data': structured_data,
            'todo_items': todo_items,
            'timeline_data': timeline_data,
            'language': minutes.language,
            'processing_status': minutes.processing_status,
            'created_at': minutes.created_at.isoformat() if minutes.created_at else None,
            'updated_at': minutes.updated_at.isoformat() if minutes.updated_at else None
        }
        
        return jsonify({
            'success': True,
            'data': data
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'获取会议纪要详情失败: {str(e)}'
        }), 500

@bp.route('/api/meeting-minutes/<int:minutes_id>/export')
@login_required
def export_meeting_minutes(minutes_id):
    """导出会议纪要"""
    try:
        minutes = MeetingMinutes.query.filter_by(id=minutes_id, user_id=current_user.id).first()
        
        if not minutes:
            return jsonify({'success': False, 'message': '会议纪要不存在或无权访问'}), 404
        
        format_type = request.args.get('format', 'json')
        
        # 重新处理数据（或从数据库加载）
        assistant = MeetingMinutesAssistant()
        
        # 构建结果对象
        result = {
            'text_processing': {
                'raw_text_sample': minutes.original_text,
                'quality_report': {}
            },
            'summary': {
                'summary_text': minutes.summary,
                'structured_data': json.loads(minutes.structured_data) if minutes.structured_data else {}
            },
            'todo_items': json.loads(minutes.todo_items) if minutes.todo_items else [],
            'timeline': {
                'data': json.loads(minutes.timeline_data) if minutes.timeline_data else []
            },
            'processing_time': minutes.updated_at.isoformat() if minutes.updated_at else None,
            'language': minutes.language
        }
        
        # 导出
        export_content = assistant.export_results(result, format_type)
        
        # 生成导出文件
        from flask import current_app
        export_dir = os.path.join(current_app.config['UPLOAD_FOLDER'], 'exports', 'meeting_minutes')
        os.makedirs(export_dir, exist_ok=True)
        
        export_filename = f"meeting_minutes_{minutes_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{format_type}"
        export_path = os.path.join(export_dir, export_filename)
        
        with open(export_path, 'w', encoding='utf-8') as f:
            f.write(export_content)
        
        return jsonify({
            'success': True,
            'message': '导出成功',
            'data': {
                'export_file': export_filename,
                'format': format_type,
                'download_url': f'/static/uploads/exports/meeting_minutes/{export_filename}'
            }
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'导出失败: {str(e)}'
        }), 500
