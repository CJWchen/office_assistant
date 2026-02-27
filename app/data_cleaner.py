"""
数据清洗与AI分析模块
提供Excel数据清洗、质量分析和导出功能
"""

import pandas as pd
import numpy as np
import json
import os
from typing import Dict, List, Any, Optional, Tuple
import logging

logger = logging.getLogger(__name__)

class DataCleaner:
    """数据清洗器"""
    
    def __init__(self, file_path: str):
        """
        初始化数据清洗器
        
        Args:
            file_path: Excel文件路径
        """
        self.file_path = file_path
        self.df = None
        self.cleaned_df = None
        self.cleaning_report = {}
        
    def load_data(self) -> bool:
        """
        加载Excel文件数据
        
        Returns:
            bool: 是否成功加载
        """
        try:
            _, ext = os.path.splitext(self.file_path)
            ext = ext.lower()
            
            if ext == '.csv':
                self.df = pd.read_csv(self.file_path)
            elif ext in ['.xlsx', '.xls']:
                # 使用openpyxl读取，支持合并单元格处理
                import openpyxl
                from openpyxl.utils import get_column_letter
                
                wb = openpyxl.load_workbook(self.file_path, data_only=True)
                ws = wb.active
                
                # 处理合并单元格
                merged_ranges = ws.merged_cells.ranges
                for merged_range in list(merged_ranges):
                    top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    top_left_value = top_left_cell.value
                    
                    for row in range(merged_range.min_row, merged_range.max_row + 1):
                        for col in range(merged_range.min_col, merged_range.max_col + 1):
                            cell = ws.cell(row=row, column=col)
                            cell.value = top_left_value
                
                data = ws.values
                columns = next(data)
                self.df = pd.DataFrame(data, columns=columns)
            else:
                logger.error(f"不支持的文件格式: {ext}")
                return False
            
            # 删除完全空的行和列
            original_shape = self.df.shape
            self.df = self.df.dropna(how='all')
            self.df = self.df.dropna(axis=1, how='all')
            
            logger.info(f"数据加载成功: {original_shape} -> {self.df.shape}")
            return True
            
        except Exception as e:
            logger.error(f"数据加载失败: {str(e)}")
            return False
    
    def analyze_data_quality(self) -> Dict[str, Any]:
        """
        分析数据质量
        
        Returns:
            Dict: 数据质量报告
        """
        if self.df is None:
            return {"error": "数据未加载"}
        
        report = {
            "basic_stats": {
                "row_count": len(self.df),
                "column_count": len(self.df.columns),
                "columns": list(self.df.columns),
                "data_types": {str(dtype): int(count) for dtype, count in self.df.dtypes.value_counts().items()}
            },
            "missing_values": {
                "total": int(self.df.isnull().sum().sum()),
                "by_column": {col: int(count) for col, count in self.df.isnull().sum().items()},
                "percentage": float(self.df.isnull().sum().sum() / (len(self.df) * len(self.df.columns)) * 100) if len(self.df) > 0 else 0
            },
            "duplicate_rows": {
                "count": int(self.df.duplicated().sum()),
                "percentage": float(self.df.duplicated().sum() / len(self.df) * 100) if len(self.df) > 0 else 0
            },
            "data_type_issues": [],
            "outlier_detection": {}
        }
        
        # 检测数据类型问题
        for col in self.df.columns:
            col_type = str(self.df[col].dtype)
            if col_type == 'object':
                # 检查是否可以转换为数值
                try:
                    pd.to_numeric(self.df[col].dropna())
                    report["data_type_issues"].append({
                        "column": col,
                        "issue": "文本列包含数值数据",
                        "suggestion": "转换为数值类型"
                    })
                except:
                    pass
        
        # 数值列的异常值检测
        numeric_cols = self.df.select_dtypes(include=['number']).columns
        if len(numeric_cols) > 0:
            outlier_info = {}
            for col in numeric_cols:
                q1 = self.df[col].quantile(0.25)
                q3 = self.df[col].quantile(0.75)
                iqr = q3 - q1
                lower_bound = q1 - 1.5 * iqr
                upper_bound = q3 + 1.5 * iqr
                
                outliers = self.df[(self.df[col] < lower_bound) | (self.df[col] > upper_bound)][col]
                outlier_count = len(outliers)
                
                if outlier_count > 0:
                    outlier_info[col] = {
                        "count": outlier_count,
                        "percentage": float(outlier_count / len(self.df) * 100),
                        "lower_bound": float(lower_bound),
                        "upper_bound": float(upper_bound),
                        "min": float(self.df[col].min()),
                        "max": float(self.df[col].max())
                    }
            
            report["outlier_detection"] = outlier_info
        
        return report
    
    def clean_data(self, options: Dict[str, Any]) -> Tuple[pd.DataFrame, Dict[str, Any]]:
        """
        根据选项清洗数据
        
        Args:
            options: 清洗选项字典
            
        Returns:
            Tuple: (清洗后的DataFrame, 清洗报告)
        """
        if self.df is None:
            raise ValueError("数据未加载，请先调用load_data()")
        
        self.cleaning_report = {
            "original_shape": self.df.shape,
            "applied_operations": [],
            "removed_rows": 0,
            "removed_columns": 0,
            "filled_missing": 0,
            "converted_columns": []
        }
        
        df_clean = self.df.copy()
        
        # 1. 缺失值处理
        missing_strategy = options.get("missing_strategy", "drop")
        
        if missing_strategy == "drop":
            # 删除包含缺失值的行
            original_rows = len(df_clean)
            df_clean = df_clean.dropna()
            removed = original_rows - len(df_clean)
            self.cleaning_report["applied_operations"].append(f"删除包含缺失值的行: {removed}行")
            self.cleaning_report["removed_rows"] += removed
            
        elif missing_strategy == "mean":
            # 数值列用均值填充
            numeric_cols = df_clean.select_dtypes(include=['number']).columns
            for col in numeric_cols:
                fill_value = df_clean[col].mean()
                missing_count = df_clean[col].isnull().sum()
                df_clean[col] = df_clean[col].fillna(fill_value)
                self.cleaning_report["filled_missing"] += missing_count
            
            self.cleaning_report["applied_operations"].append("数值列缺失值用均值填充")
            
        elif missing_strategy == "median":
            # 数值列用中位数填充
            numeric_cols = df_clean.select_dtypes(include=['number']).columns
            for col in numeric_cols:
                fill_value = df_clean[col].median()
                missing_count = df_clean[col].isnull().sum()
                df_clean[col] = df_clean[col].fillna(fill_value)
                self.cleaning_report["filled_missing"] += missing_count
            
            self.cleaning_report["applied_operations"].append("数值列缺失值用中位数填充")
            
        elif missing_strategy == "custom":
            # 自定义填充值
            fill_values = options.get("fill_values", {})
            for col, value in fill_values.items():
                if col in df_clean.columns:
                    missing_count = df_clean[col].isnull().sum()
                    df_clean[col] = df_clean[col].fillna(value)
                    self.cleaning_report["filled_missing"] += missing_count
            
            self.cleaning_report["applied_operations"].append(f"使用自定义值填充缺失值: {list(fill_values.keys())}")
        
        # 2. 重复值处理
        if options.get("remove_duplicates", True):
            original_rows = len(df_clean)
            df_clean = df_clean.drop_duplicates()
            removed = original_rows - len(df_clean)
            self.cleaning_report["applied_operations"].append(f"删除重复行: {removed}行")
            self.cleaning_report["removed_rows"] += removed
        
        # 3. 数据类型转换
        type_conversions = options.get("type_conversions", {})
        for col, target_type in type_conversions.items():
            if col in df_clean.columns:
                original_type = str(df_clean[col].dtype)
                
                try:
                    if target_type == "numeric":
                        df_clean[col] = pd.to_numeric(df_clean[col], errors='coerce')
                    elif target_type == "datetime":
                        df_clean[col] = pd.to_datetime(df_clean[col], errors='coerce')
                    elif target_type == "category":
                        df_clean[col] = df_clean[col].astype('category')
                    
                    self.cleaning_report["converted_columns"].append({
                        "column": col,
                        "from": original_type,
                        "to": target_type
                    })
                except Exception as e:
                    logger.warning(f"列 {col} 类型转换失败: {str(e)}")
        
        # 4. 异常值处理
        outlier_strategy = options.get("outlier_strategy", None)
        if outlier_strategy == "cap":
            # 使用分位数缩尾法处理异常值
            numeric_cols = df_clean.select_dtypes(include=['number']).columns
            for col in numeric_cols:
                lower = df_clean[col].quantile(0.01)
                upper = df_clean[col].quantile(0.99)
                df_clean[col] = df_clean[col].clip(lower, upper)
            
            self.cleaning_report["applied_operations"].append("使用分位数缩尾法处理异常值")
        
        elif outlier_strategy == "remove":
            # 删除异常值
            numeric_cols = df_clean.select_dtypes(include=['number']).columns
            for col in numeric_cols:
                q1 = df_clean[col].quantile(0.25)
                q3 = df_clean[col].quantile(0.75)
                iqr = q3 - q1
                lower_bound = q1 - 1.5 * iqr
                upper_bound = q3 + 1.5 * iqr
                
                mask = (df_clean[col] >= lower_bound) & (df_clean[col] <= upper_bound)
                removed = len(df_clean) - mask.sum()
                df_clean = df_clean[mask]
                self.cleaning_report["removed_rows"] += removed
            
            self.cleaning_report["applied_operations"].append("删除异常值行")
        
        # 更新最终报告
        self.cleaning_report["final_shape"] = df_clean.shape
        self.cleaning_report["cleaning_summary"] = {
            "原始数据行数": self.cleaning_report["original_shape"][0],
            "原始数据列数": self.cleaning_report["original_shape"][1],
            "清洗后行数": df_clean.shape[0],
            "清洗后列数": df_clean.shape[1],
            "总删除行数": self.cleaning_report["removed_rows"],
            "填充缺失值数": self.cleaning_report["filled_missing"],
            "转换列数": len(self.cleaning_report["converted_columns"])
        }
        
        self.cleaned_df = df_clean
        return df_clean, self.cleaning_report
    
    def export_data(self, format_type: str, output_path: Optional[str] = None) -> str:
        """
        导出清洗后的数据
        
        Args:
            format_type: 导出格式 (csv, excel, json, html)
            output_path: 输出路径（可选）
            
        Returns:
            str: 导出文件路径
        """
        if self.cleaned_df is None:
            raise ValueError("没有清洗后的数据，请先调用clean_data()")
        
        if output_path is None:
            # 生成默认输出路径
            base_name = os.path.splitext(os.path.basename(self.file_path))[0]
            output_dir = os.path.join(os.path.dirname(self.file_path), "cleaned")
            os.makedirs(output_dir, exist_ok=True)
            output_path = os.path.join(output_dir, f"{base_name}_cleaned.{format_type}")
        
        if format_type == "csv":
            self.cleaned_df.to_csv(output_path, index=False)
        elif format_type == "excel":
            self.cleaned_df.to_excel(output_path, index=False)
        elif format_type == "json":
            self.cleaned_df.to_json(output_path, orient="records", indent=2)
        elif format_type == "html":
            self.cleaned_df.to_html(output_path, index=False)
        else:
            raise ValueError(f"不支持的导出格式: {format_type}")
        
        logger.info(f"数据已导出到: {output_path} ({format_type.upper()}格式)")
        return output_path
    
    def generate_visualization_data(self) -> Dict[str, Any]:
        """
        生成可视化所需数据
        
        Returns:
            Dict: 包含图表数据的字典
        """
        if self.cleaned_df is None:
            raise ValueError("没有清洗后的数据")
        
        viz_data = {
            "column_distributions": {},
            "missing_values_comparison": {},
            "data_types_comparison": {}
        }
        
        # 获取数值列的分布数据
        numeric_cols = self.cleaned_df.select_dtypes(include=['number']).columns
        for col in numeric_cols[:5]:  # 最多5个数值列
            col_data = self.cleaned_df[col].dropna().tolist()
            viz_data["column_distributions"][col] = {
                "data": col_data,
                "summary": {
                    "min": float(self.cleaned_df[col].min()),
                    "max": float(self.cleaned_df[col].max()),
                    "mean": float(self.cleaned_df[col].mean()),
                    "median": float(self.cleaned_df[col].median())
                }
            }
        
        # 获取分类列的分布数据
        categorical_cols = self.cleaned_df.select_dtypes(include=['object', 'category']).columns
        for col in categorical_cols[:3]:  # 最多3个分类列
            value_counts = self.cleaned_df[col].value_counts().head(10)  # 前10个类别
            viz_data["column_distributions"][col] = {
                "categories": value_counts.index.tolist(),
                "counts": value_counts.values.tolist()
            }
        
        return viz_data